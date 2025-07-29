from scipy.io import loadmat, savemat
from skimage import metrics
import openpyxl
import math
def calculate_average(lst):
    # 使用sum()函数计算列表中所有元素的和
    total_sum = sum(lst)
    # 使用len()函数获取列表的长度
    length = len(lst)
    # 计算平均值
    average = total_sum / length
    return average
PSNR_EPOCH= []
SSIM_EPOCH = []
for epoch in range(1,36):
    PSNR = []
    SSIM= []
    for i in range(1,6):
        test_path = r'/root/dataset/dataset3090/test/2倍训练与测试/2test/k21-2-'+ str(i) + '.mat'
        # test_path = r'/home/dataset/T1/4/T1-4-'+ str(i) + '.mat'
        # test_path = r'/home/dataset/adult/4/adult-4-'+ str(i) + '.mat'
        pred_path = r'/root/VS-CODE/CFAT-congtouxie/matall-mat/pred' + str(epoch) + '-' + str(i) + '.mat'
        # pred_path = r'/home/code/CFAT/pred-adult-3x-new-mat/pred' + str(epoch) + '-' + str(i) + '.mat'
        # pred_path = r'/home/code/CFAT/pred-adult/pred' + str(epoch) + '-' + str(i) + '.mat'
        # pred_path = r'/home/code/CFAT/pred-mat/pred' + str(epoch) + '-' + str(i) + '.mat'

        # pred_path = r'D:\Desktop\Testdata\test\pred20' + '-' + str(i) + '.mat'

        test_data = loadmat(test_path)['img']
        pred_data = loadmat(pred_path)['data']

        psnr_value = metrics.peak_signal_noise_ratio(test_data, pred_data,data_range=1)
        PSNR.append(psnr_value)
        ssim_value = metrics.structural_similarity(test_data, pred_data,data_range=1)
        SSIM.append(ssim_value)
        print('--{}与{}--PSNR值:{}-----{}'.format(test_path, pred_path, psnr_value, ssim_value))

    PSNR_EPOCH.append(calculate_average(PSNR))
    SSIM_EPOCH.append(calculate_average(SSIM))
    print("eopch:{}五个头PSNR平均值：{}".format(epoch,calculate_average(PSNR)))
    print("eopch:{}五个头SSIM平均值：{}".format(epoch,calculate_average(SSIM)))

print('------------数据存进excel表中-----------------')
wb = openpyxl.Workbook()
ws = wb.active  # 获取活动的工作表，默认是第一个工作表\
# 写入第一个数组到第一列
for i, value in enumerate(PSNR_EPOCH, start=1):
    ws.cell(row=i, column=1, value=value)
# 写入第二个数组到第二列
for i, value in enumerate(SSIM_EPOCH, start=1):
    ws.cell(row=i, column=2, value=value)
wb.save('psnr-ssim-4-2-2.xlsx')
print('excel表保存成功！！！')

best_psnr = max(PSNR_EPOCH)
index = PSNR_EPOCH.index(best_psnr)
print("最佳epoch:{}，对应PSNR值为：{}，ssim为{}".format(index+1,best_psnr,SSIM_EPOCH[index]))





